# -*- coding: utf-8 -*-
"""
    This module defines the polygons.
    It requires modules: crystal_class, basis, CrystalLattice,
    ps_setd_mingled, mlab, tvtk and enclose_polyhedron.
"""
import numpy as np
import utils.basis as basis
from utils.crystal_lattice import CrystalLattice
from utils.cgeometry import enclosure
import utils.crystal_class as CC
from utils.crystal_shape import ps_setd_mingled, enclose_polyhedron
import random
import openpyxl
import pandas as pd
import os
import csv
from itertools import combinations
from math import factorial


def delete_xlsx_files(folder_path):
    if not os.path.exists(folder_path):
        print(f"Folder does not exist: {folder_path}")
        return

    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(folder_path, filename)
            try:
                os.remove(file_path)
                print(f"The file have been removed: {file_path}")
            except Exception as e:
                print(f"Error when removing file: {file_path} - {e}")


def crystal_plane(hkl):
    '''
       return the geometric planes of the polyhedron, where the 32 point group is considered,
	   and the lattice parameter is taken as 0.4nm.
	   hkl: Miller indices of crystallographic planes.
    '''
    cc1 = CC.CrystalClass30()  # class   (modifiable parameter)
    latt = CrystalLattice(b1, cc1)
    planes = []
    number_planes = []
    for index in hkl:
        pf = latt.geometric_plane_family(index)
        number_planes.append(len(pf))
        planes.extend(pf)
    return planes, number_planes


def surface_areas(d):
    '''
    input:
        origin-to-planes distances.

	    volume_calc: the natural or unscaled volume of the polyhedron.
		area_p: areas of all the planes.
		points_p: interception points on the planes, grouped by plane.

	return:
		area_ps: the area of each plane family (sum of areas of surfaces of a plane family).
    '''
    ps_setd_mingled(number_planes, planes, d)
    planes_c = [p.as_list() for p in planes]
    pr = enclosure(planes_c)
    volume_calc = pr[0]  # the natural or unscaled volume of the polyhedron.
    area_p = pr[1]  # areas of all the planes.
    if all(x == 0 for x in area_p):
        return area_p
    points_p = pr[2]  # interception points on the planes, grouped by plane.
    area_ps = []  # area of each plane family: sum of areas of surfaces of a plane family.
    index_start = 0
    for subset_n in number_planes:
        index_end = index_start + subset_n
        area_ps.append(np.sum(area_p[index_start: index_end]))
        index_start = index_end
    diameter = 10.0
    volume = 4.0 / 3.0 * np.pi * (diameter * 0.5) ** 3
    area_ps = np.array(area_ps) * np.power(volume / volume_calc, 2.0 / 3)

    return area_ps


if __name__ == "__main__":
    hkl_all = [[1, 3, -1], [2, -2, 2], [2, 0, 2], [2, 0, -1], [0, 4, 0]]  # HKL (modifiable parameter)
    b1 = basis.cubic(7.3253)  # Unit Cell Parameters (modifiable parameter)
    hkl_size = 4
    point_group = "432"  # Point Group (modifiable parameter)
    crystal_class = "cubic"  # Crystal System (modifiable parameter)
    output_folder = "database/"+point_group + "/" + str(hkl_size)

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
        print(f"Folder '{output_folder}' has been created successfully!")
    else:
        print(f"Folder '{output_folder}' existed.")

    for i in range(hkl_size, hkl_size + 1):
        combination_count = factorial(len(hkl_all)) // (factorial(i) * factorial(len(hkl_all) - i))
        print('===============================')
        print("There are " + str(combination_count) + " possible combinations for " + str(i) + " elements.")

        random_combinations = random.sample(list(combinations(hkl_all, i)),
                                            min(1000, len(list(combinations(hkl_all, i)))))
        random_combinations = [list(combination) for combination in random_combinations]

        # Set data_size based on the value of i
        data_size = 10000 if i in {2, 3, 5} else 2 if i == 4 else 0

        for combination in random_combinations:
            hkl = combination
            print('hkl', hkl)

            sorted_combination = sorted(combination, key=lambda x: sum(x))
            print('sorted_combination', sorted_combination)
            hkl = sorted_combination
            hkl_str = "".join([str(item) for sublist in hkl for item in sublist])
            sss = random.uniform(0.0, 2.0)
            excel_filename = "%s_%s_%s.csv" % (crystal_class, point_group, hkl_str)

            output_filename = os.path.join(output_folder, excel_filename)
            if os.path.exists(output_filename):
                print(f"File {output_filename} existed.")
                continue

            diameter = 10.0
            volume = 4.0 / 3.0 * np.pi * (diameter * 0.5) ** 3
            print(volume)
            planes, number_planes = crystal_plane(hkl)
            wb = openpyxl.Workbook()
            ws1 = wb.create_sheet("Surface_energy")
            ws2 = wb.create_sheet("Surface_area")

            gama_labels = []
            for iii, h in enumerate(hkl):
                energy_label = 'energy' + ''.join(str(x) for x in h)
                ws1.cell(row=1, column=iii + 1).value = energy_label
                area_label = 'area' + ''.join(str(x) for x in h)
                ws2.cell(row=1, column=iii + 1).value = area_label
                gama_label = 'gama' + ''.join(str(x) for x in h)

            num_hkl = len(hkl)
            k = 0

            for num in range(2000000):
                if num % 100000 == 0:
                    print("Progress ", num)
                qfp = []
                gama = []
                for j in range(num_hkl):
                    if j == 0:
                        gama.append(1)
                    else:
                        gama.append(random.uniform(0.0, 2.0))

                areas = surface_areas(gama)
                if num == 50000 and k == 0:
                    break
                if all(x == 0 for x in areas):
                    continue
                areas_fraction = areas / sum(areas)
                if k >= data_size:
                    break
                if 0.0 not in areas_fraction:  # and sum(areas_fraction)== Crystal_function:
                    # if 0.0 not in areas_fraction and sum(areas_fraction)== Crystal_function and round(areas_fraction[0],Crystal_function) != 0.5:
                    surf_dist = gama
                    ps_setd_mingled(number_planes, planes, surf_dist)
                    qfp = enclose_polyhedron(number_planes, planes, volume)
                    for j in range(num_hkl):
                        ws1.cell(row=k + 2, column=j + 1).value = gama[j]
                        ws2.cell(row=k + 2, column=j + 1).value = areas_fraction[j]

                    k = k + 1
                    print("k:", k)
            wb.save(output_filename)

            xlsx = pd.ExcelFile(output_filename)
            df_energy = pd.read_excel(xlsx, 'Surface_energy')
            df_area = pd.read_excel(xlsx, 'Surface_area')

            df_all = pd.concat([df_energy, df_area], axis=1)

            scv_filename = "%s_%s_%s.csv" % (crystal_class, point_group, hkl_str)
            output_filename = os.path.join(output_folder, scv_filename)

            df_all.to_csv(output_filename, sep=',', index=False)

            model_name = "%s_%s" % (point_group, hkl_str)
            data_to_write = [model_name, crystal_class, point_group]

            with open(output_filename, 'r', newline='') as file:
                reader = csv.reader(file)
                rows = list(reader)

            with open(output_filename, 'w', newline='') as file:
                writer = csv.writer(file)
                writer.writerows(rows)

            folder_path = output_folder
            delete_xlsx_files(folder_path)
